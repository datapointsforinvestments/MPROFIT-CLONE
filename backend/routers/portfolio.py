"""
portfolio.py — Multi-folio portfolio tracker router.
Endpoints for folios, transactions, bulk upload, summary, P&L, reconciliation,
symbol mappings, and live price refresh.
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime, timedelta
from typing import Optional

import pandas as pd
import yfinance as yf
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session

from database import get_db
from models.portfolio import (
    Folio, PortfolioAsset, PortfolioQuote, PortfolioSymbolMapping, PortfolioTransaction, PortfolioDividend,
)
from models.benchmark import BenchmarkIndex
from routers.auth import get_current_user, require_admin, require_fm_or_above
from services import financial as fin
from services import name_mapper as nm

router = APIRouter(prefix="/api/portfolio", tags=["portfolio"])
logger = logging.getLogger(__name__)

TYPE_MAP = {
    "buy": "Buy", "sell": "Sell", "bonus": "Bonus", "split": "Split",
    "dividend": "Dividend", "transfer in": "Transfer_In", "transfer out": "Transfer_Out",
}
EXPECTED_COLS = {"Date", "Trans. Type", "Asset", "Qty", "Price", "Amount", "Folio"}


# ── helpers ───────────────────────────────────────────────────────────────────

def _enrich_asset(symbol: str) -> dict:
    """Fetch name/sector/industry from yfinance for a new portfolio asset."""
    try:
        t = yf.Ticker(f"{symbol}.NS")
        info = t.info or {}
        name = (info.get("longName") or info.get("shortName") or symbol)
        sector = info.get("sector") or "Unknown"
        industry = info.get("industry") or ""
        if name == symbol:
            t2 = yf.Ticker(f"{symbol}.BO")
            info2 = t2.info or {}
            name = info2.get("longName") or info2.get("shortName") or symbol
            sector = info2.get("sector") or sector
            industry = info2.get("industry") or industry
    except Exception:
        name = symbol
        sector = "Unknown"
        industry = ""
    return {"name": name, "sector": sector, "industry": industry}


def _to_txn_dicts(txns: list) -> list:
    return [
        {
            "trade_date": t.trade_date,
            "trans_type": t.trans_type,
            "quantity": float(t.quantity),
            "price": float(t.price),
            "total_amount": float(t.total_amount),
            "split_ratio": float(t.split_ratio) if t.split_ratio else None,
        }
        for t in txns
    ]


# ── Folios ────────────────────────────────────────────────────────────────────

@router.get("/folios")
def list_folios(db: Session = Depends(get_db), _=Depends(get_current_user)):
    folios = db.query(Folio).filter(Folio.is_active == True).order_by(Folio.name).all()
    return [{"id": f.id, "name": f.name, "is_active": f.is_active} for f in folios]


@router.delete("/folios/{folio_id}", status_code=204)
def delete_folio(folio_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    folio = db.query(Folio).filter(Folio.id == folio_id).first()
    if not folio:
        raise HTTPException(404, "Folio not found")
    has_txns = db.query(PortfolioTransaction).filter(PortfolioTransaction.folio_id == folio_id).first()
    if has_txns:
        raise HTTPException(409, f"Folio '{folio.name}' has transactions. Delete all transactions first.")
    db.delete(folio)
    db.commit()


@router.post("/folios", status_code=201)
def create_folio(body: dict, db: Session = Depends(get_db), _=Depends(require_fm_or_above)):
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "name is required")
    existing = db.query(Folio).filter(Folio.name.ilike(name)).first()
    if existing:
        raise HTTPException(409, f"Folio '{name}' already exists")
    folio = Folio(name=name)
    db.add(folio)
    db.commit()
    db.refresh(folio)
    return {"id": folio.id, "name": folio.name, "is_active": folio.is_active}


# ── Upload Template ──────────────────────────────────────────────────────────

@router.get("/upload-template")
def download_upload_template(_=Depends(get_current_user)):
    """Return a sample Excel file with the correct column headers."""
    import openpyxl
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = ["Date", "Trans. Type", "Asset", "Qty", "Price", "Amount", "Folio"]
    ws.append(headers)
    # Sample rows
    ws.append(["01-01-2024", "buy",  "RELIANCE", 10, 2500, 25000, "Main"])
    ws.append(["15-03-2024", "sell", "INFY",     5,  1800, 9000,  "Main"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=portfolio_template.xlsx"},
    )


# ── Bulk Upload ───────────────────────────────────────────────────────────────

@router.post("/upload")
def upload_transactions(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_fm_or_above),
):
    content = file.file.read()
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    try:
        if ext in ("xlsx", "xls"):
            xls = pd.ExcelFile(io.BytesIO(content))
            required_norm = {c.lower().replace(".", "").replace(" ", "") for c in EXPECTED_COLS}
            matching = []
            for sname in xls.sheet_names:
                try:
                    candidate = xls.parse(sname)
                    candidate.columns = [" ".join(str(c).strip().split()) for c in candidate.columns]
                    norm = {c.lower().replace(".", "").replace(" ", "") for c in candidate.columns}
                    if required_norm.issubset(norm):
                        matching.append(candidate)
                except Exception:
                    continue
            df = pd.concat(matching, ignore_index=True) if matching else xls.parse(0)
        elif ext == "csv":
            df = pd.read_csv(io.BytesIO(content))
        else:
            raise HTTPException(400, "Only .xlsx/.xls/.csv accepted")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Cannot parse file: {e}")

    # Normalise column names: strip whitespace, collapse multiple spaces
    df.columns = [" ".join(str(c).strip().split()) for c in df.columns]
    # Build a case-insensitive map from actual column → canonical name
    col_norm = {c.lower().replace(".", "").replace(" ", ""): c for c in df.columns}
    canonical_map = {}
    for expected in EXPECTED_COLS:
        key = expected.lower().replace(".", "").replace(" ", "")
        if expected in df.columns:
            canonical_map[expected] = expected
        elif key in col_norm:
            canonical_map[expected] = col_norm[key]
    missing = EXPECTED_COLS - set(canonical_map.keys())
    if missing:
        raise HTTPException(400, f"Missing columns: {missing}. Found: {list(df.columns)}")
    # Rename columns to canonical names
    df = df.rename(columns={v: k for k, v in canonical_map.items() if k != v})

    # Build lookup caches
    folio_map = {f.name.strip().lower(): f.id for f in db.query(Folio).all()}
    all_assets = db.query(PortfolioAsset).all()
    asset_map = {a.symbol.upper(): a.id for a in all_assets}

    custom_map = {m.norm_name: m.symbol for m in db.query(PortfolioSymbolMapping).all()}
    dynamic_map = nm.build_dynamic_map([{"symbol": a.symbol, "name": a.name} for a in all_assets])

    # Pre-create all new PortfolioAsset rows WITHOUT yfinance (use symbol as name placeholder).
    # This avoids N × yfinance round-trips inside the row loop.
    # A background refresh-quotes call will enrich names/sectors afterwards.
    unique_raw_assets = df["Asset"].dropna().unique()
    for raw in unique_raw_assets:
        sym = nm.resolve_symbol(str(raw).strip(), custom_map, dynamic_map)
        if sym and sym.upper() not in asset_map:
            na = PortfolioAsset(symbol=sym.upper(), name=sym.upper())
            db.add(na)
            db.flush()
            asset_map[sym.upper()] = na.id
            dynamic_map[nm.normalize(sym)] = sym.upper()

    existing_set: set[tuple] = set(
        (r.folio_id, r.asset_id, r.trade_date, r.trans_type, float(r.quantity), float(r.price))
        for r in db.query(PortfolioTransaction).all()
    )

    rows_ok = rows_fail = rows_duplicate = 0
    errors: list[str] = []
    unmapped_names: list[str] = []
    pending_txns: list[dict] = []

    for i, row in df.iterrows():
        lineno = int(i) + 2
        try:
            raw_folio = str(row["Folio"]).strip()
            raw_asset = str(row["Asset"]).strip()
            raw_type  = str(row["Trans. Type"]).strip().lower()

            resolved = nm.resolve_symbol(raw_asset, custom_map, dynamic_map)
            if resolved is None:
                if raw_asset not in unmapped_names:
                    unmapped_names.append(raw_asset)
                rows_fail += 1
                errors.append(f"Row {lineno}: Cannot resolve '{raw_asset}' — add a custom mapping.")
                continue

            symbol = resolved.upper()
            def _num(v):
                try:
                    f = float(str(v).replace(",", "") or 0)
                    return 0.0 if (f != f) else f
                except Exception:
                    return 0.0
            qty    = _num(row["Qty"])
            price  = _num(row["Price"])
            amount = _num(row["Amount"]) or qty * price

            raw_date = row["Date"]
            if isinstance(raw_date, str):
                trade_date = None
                for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y", "%b %d %Y"):
                    try:
                        trade_date = datetime.strptime(raw_date.strip(), fmt).date()
                        break
                    except ValueError:
                        pass
                if not trade_date:
                    raise ValueError(f"Unrecognised date: {raw_date}")
            elif isinstance(raw_date, datetime):
                trade_date = raw_date.date()
            elif hasattr(raw_date, "date"):
                trade_date = raw_date.date()
            elif isinstance(raw_date, date):
                trade_date = raw_date
            else:
                raise ValueError(f"Bad date type: {type(raw_date)}")

            trans_type = TYPE_MAP.get(raw_type)
            if not trans_type:
                raise ValueError(f"Unknown transaction type: '{raw_type}'")

            # Find-or-create folio
            folio_id = folio_map.get(raw_folio.lower())
            if not folio_id:
                nf = Folio(name=raw_folio)
                db.add(nf)
                db.flush()
                folio_map[raw_folio.lower()] = nf.id
                folio_id = nf.id

            # Asset was pre-created above; just look it up
            asset_id = asset_map.get(symbol)

            # Duplicate check
            dup_key = (folio_id, asset_id, trade_date, trans_type, qty, price)
            if dup_key in existing_set:
                rows_duplicate += 1
                continue
            existing_set.add(dup_key)

            pending_txns.append(dict(
                folio_id=folio_id, asset_id=asset_id, trade_date=trade_date,
                trans_type=trans_type, quantity=qty, price=price, total_amount=amount,
                brokerage=0, notes=None, created_by=current_user.username,
            ))
            rows_ok += 1

        except Exception as e:
            rows_fail += 1
            errors.append(f"Row {lineno}: {e}")

    # Bulk insert with ON CONFLICT DO NOTHING — safe against any partial prior runs
    from sqlalchemy.dialects.postgresql import insert as pg_insert
    rows_inserted = 0
    if pending_txns:
        stmt = pg_insert(PortfolioTransaction.__table__).values(pending_txns).on_conflict_do_nothing()
        result = db.execute(stmt)
        rows_inserted = result.rowcount
        rows_duplicate += (rows_ok - rows_inserted)
        rows_ok = rows_inserted
    db.commit()
    return {
        "rows_processed": len(df),
        "rows_successful": rows_ok,
        "rows_failed": rows_fail,
        "rows_duplicate": rows_duplicate,
        "errors": errors[:50],
        "unmapped_names": unmapped_names[:50],
    }


# ── Portfolio Summary ─────────────────────────────────────────────────────────

@router.get("/summary")
def portfolio_summary(
    folio_id: Optional[int] = Query(None),
    consolidated: bool = Query(False),
    include_exited: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    if folio_id:
        folios = db.query(Folio).filter(Folio.id == folio_id).all()
    else:
        folios = db.query(Folio).filter(Folio.is_active == True).order_by(Folio.name).all()
    if not folios:
        raise HTTPException(404, "No folios found")

    asset_map = {a.id: a for a in db.query(PortfolioAsset).all()}
    quote_map = {q.asset_id: q for q in db.query(PortfolioQuote).all()}

    folio_ids = [f.id for f in folios]
    all_txns = (
        db.query(PortfolioTransaction)
        .filter(PortfolioTransaction.folio_id.in_(folio_ids))
        .order_by(PortfolioTransaction.trade_date)
        .all()
    )

    folio_summaries = []
    all_portfolio_div_cfs: list[tuple] = []  # Accumulates across folios for consolidated

    for folio in folios:
        grouped: dict[int, list] = {}
        for t in all_txns:
            if t.folio_id == folio.id:
                grouped.setdefault(t.asset_id, []).append(t)

        # Preload dividends for this folio — keyed by asset_id
        folio_div_map: dict[int, list] = {}
        for d in db.query(PortfolioDividend).filter(PortfolioDividend.folio_id == folio.id).all():
            folio_div_map.setdefault(d.asset_id, []).append(d)

        trailing_cutoff = date.today() - timedelta(days=365)

        holdings = []
        all_cfs = []
        total_current = 0.0
        total_invested_folio = 0.0

        for asset_id, txns in grouped.items():
            asset = asset_map.get(asset_id)
            if not asset:
                continue
            q   = quote_map.get(asset_id)
            cmp = float(q.cmp) if q and q.cmp else 0.0

            txn_dicts = _to_txn_dicts(txns)
            qty, avg_p, invested = fin.compute_avg_price_and_qty(txn_dicts)
            pnl = fin.fifo_pnl(txn_dicts, asset.symbol, asset.name, folio.name)
            realised = sum(r["gain_loss"] for r in pnl.rows)
            first_buy = min(
                (t.trade_date for t in txns if t.trans_type == "Buy"), default=None
            )

            if qty < 0.0001:
                if include_exited:
                    last_sell = max(
                        (t.trade_date for t in txns if t.trans_type == "Sell"), default=None
                    )
                    holdings.append({
                        "folio_id": folio.id, "folio_name": folio.name,
                        "asset_id": asset_id, "symbol": asset.symbol,
                        "asset_name": asset.name, "sector": asset.sector,
                        "net_qty": 0.0, "avg_price": round(avg_p, 4),
                        "total_investment": round(invested, 2), "cmp": None,
                        "current_value": 0.0,
                        "unrealised_pnl": None, "unrealised_pnl_pct": None,
                        "realised_pnl": round(realised, 2),
                        "realised_pnl_pct": round(realised / invested * 100, 2) if invested else 0,
                        "xirr_pct": None, "cagr_pct": None,
                        "first_purchase_date": first_buy.isoformat() if first_buy else None,
                        "last_exit_date": last_sell.isoformat() if last_sell else None,
                        "is_exited": True,
                        "day_change_pct": None,
                    })
                continue

            current_val = cmp * qty if cmp else 0.0
            cfs = fin.build_xirr_cashflows(txn_dicts, current_val)
            all_cfs.extend(cfs)
            asset_xirr = fin.xirr(cfs) if cfs else None
            asset_cagr = fin.cagr(invested, current_val, first_buy) if first_buy and invested else None
            unrealised = current_val - invested if cmp else None
            unrealised_pct = (
                round(unrealised / invested * 100, 2) if unrealised is not None and invested else None
            )

            # Dividend enrichment
            holding_divs = folio_div_map.get(asset_id, [])
            total_div = round(sum(float(d.total_received or 0) for d in holding_divs), 2)
            trailing_div = round(sum(float(d.total_received or 0) for d in holding_divs if d.ex_date >= trailing_cutoff), 2)
            div_cfs = [(d.ex_date, float(d.total_received or 0)) for d in holding_divs if (d.total_received or 0) > 0]
            div_xirr_cfs = sorted(cfs + div_cfs, key=lambda x: x[0])
            div_xirr = fin.xirr(div_xirr_cfs) if div_cfs and div_xirr_cfs else asset_xirr

            total_current        += current_val
            total_invested_folio += invested

            holdings.append({
                "folio_id": folio.id, "folio_name": folio.name,
                "asset_id": asset_id, "symbol": asset.symbol,
                "asset_name": asset.name, "sector": asset.sector,
                "net_qty": round(qty, 4), "avg_price": round(avg_p, 4),
                "total_investment": round(invested, 2),
                "cmp": round(cmp, 2) if cmp else None,
                "current_value": round(current_val, 2) if current_val else None,
                "unrealised_pnl": round(unrealised, 2) if unrealised is not None else None,
                "unrealised_pnl_pct": unrealised_pct,
                "realised_pnl": round(realised, 2),
                "realised_pnl_pct": round(realised / invested * 100, 2) if invested else 0,
                "xirr_pct": asset_xirr, "cagr_pct": asset_cagr,
                "total_dividend": total_div, "trailing_div": trailing_div, "div_xirr_pct": div_xirr,
                "first_purchase_date": first_buy.isoformat() if first_buy else None,
                "last_exit_date": None, "is_exited": False,
                "day_change_pct": float(q.day_change_pct) if q and q.day_change_pct else None,
            })

        # portfolio_pct
        if total_current > 0:
            for h in holdings:
                cv = h.get("current_value") or 0
                h["portfolio_pct"] = round(cv / total_current * 100, 2)
        else:
            for h in holdings:
                h["portfolio_pct"] = 0.0

        folio_gain = total_current - total_invested_folio
        folio_gain_pct = (folio_gain / total_invested_folio * 100) if total_invested_folio else 0.0
        folio_xirr = fin.xirr(all_cfs) if all_cfs else None
        folio_cagr = fin.cagr(total_invested_folio, total_current, min(
            (t.trade_date for t in all_txns if t.folio_id == folio.id and t.trans_type == "Buy"),
            default=None
        )) if total_invested_folio else None

        holdings.sort(key=lambda x: (x["is_exited"], -(x.get("current_value") or 0)))

        folio_total_div    = round(sum(h.get("total_dividend", 0) for h in holdings), 2)
        folio_trailing_div = round(sum(h.get("trailing_div",    0) for h in holdings), 2)

        # Portfolio-level XIRR including dividends
        folio_all_div_cfs = [
            (d.ex_date, float(d.total_received or 0))
            for divs in folio_div_map.values()
            for d in divs
            if (d.total_received or 0) > 0
        ]
        all_portfolio_div_cfs.extend(folio_all_div_cfs)
        folio_div_xirr_cfs = sorted(all_cfs + folio_all_div_cfs, key=lambda x: x[0])
        folio_div_xirr = fin.xirr(folio_div_xirr_cfs) if folio_all_div_cfs and folio_div_xirr_cfs else folio_xirr

        folio_summaries.append({
            "folio_id": folio.id, "folio_name": folio.name,
            "total_investment": round(total_invested_folio, 2),
            "current_value": round(total_current, 2),
            "total_gain": round(folio_gain, 2),
            "total_gain_pct": round(folio_gain_pct, 4),
            "xirr_pct": folio_xirr, "cagr_pct": folio_cagr,
            "div_xirr_pct": folio_div_xirr,
            "total_dividend": folio_total_div,
            "trailing_12m_dividend": folio_trailing_div,
            "holdings": holdings,
        })

    if consolidated:
        all_inv      = sum(s["total_investment"]      for s in folio_summaries)
        all_div      = round(sum(s["total_dividend"]      for s in folio_summaries), 2)
        all_trail_div= round(sum(s["trailing_12m_dividend"] for s in folio_summaries), 2)
        all_cur  = sum(s["current_value"]    for s in folio_summaries)
        all_gain = all_cur - all_inv
        all_pct  = (all_gain / all_inv * 100) if all_inv else 0.0
        cons_cfs = []
        for t in all_txns:
            d = t.trade_date
            if isinstance(d, datetime):
                d = d.date()
            if t.trans_type == "Buy":
                cons_cfs.append((d, -float(t.total_amount)))
            elif t.trans_type == "Sell":
                cons_cfs.append((d, float(t.total_amount)))
        if all_cur:
            cons_cfs.append((date.today(), all_cur))
        cons_cfs.sort(key=lambda x: x[0])
        cons_xirr = fin.xirr(cons_cfs) if cons_cfs else None
        cons_div_xirr_cfs = sorted(cons_cfs + all_portfolio_div_cfs, key=lambda x: x[0])
        cons_div_xirr = fin.xirr(cons_div_xirr_cfs) if all_portfolio_div_cfs and cons_div_xirr_cfs else cons_xirr

        # Merge holdings across folios for consolidated view
        merged: dict[str, dict] = {}
        for fs in folio_summaries:
            for h in fs["holdings"]:
                sym = h["symbol"]
                if sym not in merged:
                    merged[sym] = dict(h)
                    merged[sym]["folio_names"] = [h["folio_name"]]
                else:
                    m = merged[sym]
                    m["folio_names"].append(h["folio_name"])
                    m["net_qty"] = round(m["net_qty"] + h["net_qty"], 4)
                    m["total_investment"] = round(m["total_investment"] + h["total_investment"], 2)
                    cv1 = m.get("current_value") or 0
                    cv2 = h.get("current_value") or 0
                    m["current_value"] = round(cv1 + cv2, 2) if (cv1 or cv2) else None
                    up1 = m.get("unrealised_pnl") or 0
                    up2 = h.get("unrealised_pnl") or 0
                    m["unrealised_pnl"] = round(up1 + up2, 2) if (m.get("unrealised_pnl") is not None or h.get("unrealised_pnl") is not None) else None
                    m["realised_pnl"] = round((m.get("realised_pnl") or 0) + (h.get("realised_pnl") or 0), 2)
                    m["total_dividend"] = round((m.get("total_dividend") or 0) + (h.get("total_dividend") or 0), 2)
                    m["trailing_div"]   = round((m.get("trailing_div") or 0) + (h.get("trailing_div") or 0), 2)
                    # div_xirr: not re-computed across folios; take weighted average by current_value
                    cv_total = (cv1 or 0) + (cv2 or 0)
                    if cv_total > 0:
                        xi1 = (m.get("div_xirr_pct") or 0) * (cv1 or 0)
                        xi2 = (h.get("div_xirr_pct") or 0) * (cv2 or 0)
                        m["div_xirr_pct"] = round((xi1 + xi2) / cv_total, 4)
                    else:
                        m["div_xirr_pct"] = m.get("div_xirr_pct")

        for m in merged.values():
            m["folio_name"] = ", ".join(m["folio_names"])
            inv = m.get("total_investment") or 0
            cv  = m.get("current_value") or 0
            m["avg_price"] = round(m["total_investment"] / m["net_qty"], 4) if m["net_qty"] else 0
            m["unrealised_pnl_pct"] = round((m["unrealised_pnl"] / inv * 100), 2) if (m.get("unrealised_pnl") is not None and inv) else None
            m["portfolio_pct"] = round(cv / all_cur * 100, 2) if all_cur else 0.0

        cons_holdings = sorted(merged.values(), key=lambda x: (x["is_exited"], -(x.get("current_value") or 0)))

        return {
            "consolidated": True,
            "total_investment": round(all_inv, 2),
            "current_value": round(all_cur, 2),
            "total_gain": round(all_gain, 2),
            "total_gain_pct": round(all_pct, 4),
            "xirr_pct": cons_xirr,
            "div_xirr_pct": cons_div_xirr,
            "total_dividend": all_div,
            "trailing_12m_dividend": all_trail_div,
            "folios": folio_summaries,
            "holdings": cons_holdings,
        }

    return folio_summaries[0] if len(folio_summaries) == 1 else folio_summaries


# ── P&L Report ────────────────────────────────────────────────────────────────

@router.get("/pl-report")
def pl_report(
    folio_id:  Optional[int]  = Query(None),
    from_date: Optional[date] = Query(None),
    to_date:   Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(PortfolioTransaction)
    if folio_id:
        q = q.filter(PortfolioTransaction.folio_id == folio_id)
    all_txns = q.all()

    folio_map = {f.id: f.name for f in db.query(Folio).all()}
    asset_map = {a.id: a for a in db.query(PortfolioAsset).all()}

    grouped: dict[tuple, list] = {}
    for t in all_txns:
        grouped.setdefault((t.folio_id, t.asset_id), []).append(t)

    reports = []
    total_stcg = 0.0
    total_ltcg = 0.0
    for (fid, aid), txns in grouped.items():
        asset = asset_map.get(aid)
        folio_name = folio_map.get(fid, "Unknown")
        if not asset:
            continue
        txn_dicts = _to_txn_dicts(txns)
        r = fin.fifo_pnl(txn_dicts, asset.symbol, asset.name, folio_name)
        rows = r.rows
        if from_date:
            rows = [x for x in rows if x["sell_date"] >= from_date]
        if to_date:
            rows = [x for x in rows if x["sell_date"] <= to_date]
        if rows:
            stcg = sum(x["gain_loss"] for x in rows if x["tax_category"] == "STCG")
            ltcg = sum(x["gain_loss"] for x in rows if x["tax_category"] == "LTCG")
            total_stcg += stcg
            total_ltcg += ltcg
            reports.append({
                "folio_name": folio_name,
                "symbol": asset.symbol,
                "asset_name": asset.name,
                "realised_stcg": round(stcg, 2),
                "realised_ltcg": round(ltcg, 2),
                "total_realised": round(stcg + ltcg, 2),
                "transactions": [
                    {
                        **row,
                        "buy_date": row["buy_date"].isoformat() if hasattr(row["buy_date"], "isoformat") else str(row["buy_date"]),
                        "sell_date": row["sell_date"].isoformat() if hasattr(row["sell_date"], "isoformat") else str(row["sell_date"]),
                    }
                    for row in rows
                ],
            })

    return {
        "summary": {
            "total_stcg": round(total_stcg, 2),
            "total_ltcg": round(total_ltcg, 2),
            "total_realised": round(total_stcg + total_ltcg, 2),
        },
        "reports": sorted(reports, key=lambda x: x["total_realised"], reverse=True),
    }


# ── Asset metadata update ────────────────────────────────────────────────────

@router.put("/assets/{asset_id}")
def update_asset(
    asset_id: int,
    body: dict,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    asset = db.query(PortfolioAsset).filter(PortfolioAsset.id == asset_id).first()
    if not asset:
        raise HTTPException(404, "Asset not found")
    if "sector" in body:
        asset.sector = body["sector"] or None
    if "name" in body:
        asset.name = body["name"] or asset.symbol
    if "symbol" in body and body["symbol"]:
        new_sym = body["symbol"].strip().upper()
        conflict = db.query(PortfolioAsset).filter(
            PortfolioAsset.symbol == new_sym, PortfolioAsset.id != asset_id
        ).first()
        if conflict:
            raise HTTPException(400, f"Symbol '{new_sym}' already in use")
        asset.symbol = new_sym
        # Invalidate stale quote so next refresh fetches fresh price
        db.query(PortfolioQuote).filter(PortfolioQuote.asset_id == asset_id).delete()
    db.commit()
    return {"id": asset.id, "symbol": asset.symbol, "name": asset.name, "sector": asset.sector}


# ── Transactions ──────────────────────────────────────────────────────────────

@router.get("/transactions")
def list_transactions(
    folio_id:   Optional[int]  = Query(None),
    symbol:     Optional[str]  = Query(None),
    from_date:  Optional[date] = Query(None),
    to_date:    Optional[date] = Query(None),
    trans_type: Optional[str]  = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(PortfolioTransaction)
    if folio_id:
        q = q.filter(PortfolioTransaction.folio_id == folio_id)
    if symbol:
        asset = db.query(PortfolioAsset).filter(PortfolioAsset.symbol == symbol.upper()).first()
        if asset:
            q = q.filter(PortfolioTransaction.asset_id == asset.id)
    if from_date:
        q = q.filter(PortfolioTransaction.trade_date >= from_date)
    if to_date:
        q = q.filter(PortfolioTransaction.trade_date <= to_date)
    if trans_type:
        q = q.filter(PortfolioTransaction.trans_type == trans_type)
    txns = q.order_by(PortfolioTransaction.trade_date.desc()).all()

    folio_map = {f.id: f.name for f in db.query(Folio).all()}
    asset_map = {a.id: a for a in db.query(PortfolioAsset).all()}

    def _safe(v):
        try:
            f = float(v)
            return 0.0 if (f != f) else f  # nan check
        except Exception:
            return 0.0

    return [
        {
            "id": t.id, "folio_id": t.folio_id,
            "folio_name": folio_map.get(t.folio_id, ""),
            "asset_id": t.asset_id,
            "symbol": asset_map[t.asset_id].symbol if t.asset_id in asset_map else "",
            "asset_name": asset_map[t.asset_id].name if t.asset_id in asset_map else "",
            "trade_date": t.trade_date.isoformat(),
            "trans_type": t.trans_type,
            "quantity": _safe(t.quantity),
            "price": _safe(t.price),
            "total_amount": _safe(t.total_amount),
            "brokerage": _safe(t.brokerage or 0),
            "notes": t.notes,
            "created_by": t.created_by,
        }
        for t in txns
    ]


@router.get("/transactions/export")
def export_transactions(
    folio_id:   Optional[int]  = Query(None),
    symbol:     Optional[str]  = Query(None),
    from_date:  Optional[date] = Query(None),
    to_date:    Optional[date] = Query(None),
    trans_type: Optional[str]  = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    import openpyxl
    from fastapi.responses import StreamingResponse

    q = db.query(PortfolioTransaction)
    if folio_id:
        q = q.filter(PortfolioTransaction.folio_id == folio_id)
    if symbol:
        asset = db.query(PortfolioAsset).filter(PortfolioAsset.symbol == symbol.upper()).first()
        if asset:
            q = q.filter(PortfolioTransaction.asset_id == asset.id)
    if from_date:
        q = q.filter(PortfolioTransaction.trade_date >= from_date)
    if to_date:
        q = q.filter(PortfolioTransaction.trade_date <= to_date)
    if trans_type:
        q = q.filter(PortfolioTransaction.trans_type == trans_type)
    txns = q.order_by(PortfolioTransaction.trade_date.desc()).all()

    folio_map = {f.id: f.name for f in db.query(Folio).all()}
    asset_map = {a.id: a for a in db.query(PortfolioAsset).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = ["Date", "Folio", "Symbol", "Asset Name", "Trans. Type", "Qty", "Price", "Amount", "Brokerage", "Notes"]
    ws.append(headers)

    # Bold header
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    def _safe(v):
        try:
            f = float(v)
            return 0.0 if (f != f) else f
        except Exception:
            return 0.0

    for t in txns:
        asset = asset_map.get(t.asset_id)
        ws.append([
            t.trade_date.isoformat() if t.trade_date else "",
            folio_map.get(t.folio_id, ""),
            asset.symbol if asset else "",
            asset.name if asset else "",
            t.trans_type,
            _safe(t.quantity),
            _safe(t.price),
            _safe(t.total_amount),
            _safe(t.brokerage or 0),
            t.notes or "",
        ])

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "transactions_export.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/transactions", status_code=201)
def add_transaction(body: dict, db: Session = Depends(get_db), current_user=Depends(require_fm_or_above)):
    folio_id   = body.get("folio_id")
    symbol     = (body.get("symbol") or "").upper()
    trade_date = body.get("trade_date")
    trans_type = body.get("trans_type")
    qty        = float(body.get("quantity") or 0)
    price      = float(body.get("price") or 0)
    amount     = float(body.get("total_amount") or 0) or qty * price

    if not all([folio_id, symbol, trade_date, trans_type]):
        raise HTTPException(400, "folio_id, symbol, trade_date, trans_type are required")

    if isinstance(trade_date, str):
        try:
            trade_date = date.fromisoformat(trade_date)
        except ValueError:
            raise HTTPException(400, f"Invalid date: {trade_date}")

    folio = db.query(Folio).filter(Folio.id == folio_id).first()
    if not folio:
        raise HTTPException(404, "Folio not found")

    asset = db.query(PortfolioAsset).filter(PortfolioAsset.symbol == symbol).first()
    if not asset:
        meta = _enrich_asset(symbol)
        asset = PortfolioAsset(symbol=symbol, **meta)
        db.add(asset)
        db.flush()

    txn = PortfolioTransaction(
        folio_id=folio_id, asset_id=asset.id, trade_date=trade_date,
        trans_type=trans_type, quantity=qty, price=price, total_amount=amount,
        notes=body.get("notes"), created_by=current_user.username,
    )
    db.add(txn)
    db.commit()
    db.refresh(txn)
    return {"id": txn.id, "message": "Transaction added"}


@router.delete("/transactions/{txn_id}", status_code=204)
def delete_transaction(txn_id: int, db: Session = Depends(get_db), _=Depends(require_admin)):
    txn = db.query(PortfolioTransaction).filter(PortfolioTransaction.id == txn_id).first()
    if not txn:
        raise HTTPException(404, "Transaction not found")
    db.delete(txn)
    db.commit()


@router.put("/transactions/{txn_id}")
def update_transaction(txn_id: int, body: dict, db: Session = Depends(get_db), _=Depends(require_fm_or_above)):
    txn = db.query(PortfolioTransaction).filter(PortfolioTransaction.id == txn_id).first()
    if not txn:
        raise HTTPException(404, "Transaction not found")
    if "trade_date" in body:
        try:
            txn.trade_date = date.fromisoformat(body["trade_date"])
        except ValueError:
            raise HTTPException(400, f"Invalid date: {body['trade_date']}")
    if "trans_type" in body:
        txn.trans_type = body["trans_type"]
    if "quantity" in body:
        txn.quantity = float(body["quantity"])
    if "price" in body:
        txn.price = float(body["price"])
    txn.total_amount = float(txn.quantity) * float(txn.price)
    if "notes" in body:
        txn.notes = body.get("notes") or None
    db.commit()
    return {"id": txn.id, "message": "Updated"}


# ── Symbol Mappings ───────────────────────────────────────────────────────────

@router.get("/symbol-mappings")
def list_symbol_mappings(db: Session = Depends(get_db), _=Depends(get_current_user)):
    mappings = db.query(PortfolioSymbolMapping).order_by(PortfolioSymbolMapping.raw_name).all()
    return [{"id": m.id, "raw_name": m.raw_name, "norm_name": m.norm_name, "symbol": m.symbol} for m in mappings]


@router.post("/symbol-mappings", status_code=201)
def create_symbol_mapping(
    body: dict, db: Session = Depends(get_db), _=Depends(get_current_user)
):
    raw_name = (body.get("raw_name") or "").strip()
    symbol   = (body.get("symbol") or "").strip().upper()
    if not raw_name or not symbol:
        raise HTTPException(400, "raw_name and symbol are required")

    norm = nm.normalize(raw_name)
    existing = db.query(PortfolioSymbolMapping).filter(PortfolioSymbolMapping.norm_name == norm).first()
    if existing:
        existing.raw_name = raw_name
        existing.symbol   = symbol
        db.commit()
        db.refresh(existing)
        return {"id": existing.id, "raw_name": existing.raw_name, "norm_name": existing.norm_name, "symbol": existing.symbol}

    mapping = PortfolioSymbolMapping(raw_name=raw_name, norm_name=norm, symbol=symbol)
    db.add(mapping)
    db.commit()
    db.refresh(mapping)
    return {"id": mapping.id, "raw_name": mapping.raw_name, "norm_name": mapping.norm_name, "symbol": mapping.symbol}


@router.delete("/symbol-mappings/{mapping_id}", status_code=204)
def delete_symbol_mapping(mapping_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    m = db.query(PortfolioSymbolMapping).filter(PortfolioSymbolMapping.id == mapping_id).first()
    if not m:
        raise HTTPException(404, "Mapping not found")
    db.delete(m)
    db.commit()


# ── Demat Reconciliation ──────────────────────────────────────────────────────

@router.post("/reconcile")
def reconcile_demat(
    file: UploadFile = File(...),
    folio_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    content = file.file.read()
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    try:
        if ext in ("xlsx", "xls"):
            df = pd.read_excel(io.BytesIO(content))
        elif ext == "csv":
            df = pd.read_csv(io.BytesIO(content))
        else:
            raise HTTPException(400, "Only .xlsx/.xls/.csv accepted")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Cannot parse file: {e}")

    df.columns = [str(c).strip() for c in df.columns]
    sym_col = next((c for c in df.columns if c.lower() in ("symbol", "ticker", "isin", "scrip", "stock")), None)
    qty_col = next((c for c in df.columns if c.lower() in ("qty", "quantity", "shares", "balance", "units")), None)
    if not sym_col or not qty_col:
        raise HTTPException(400, f"Need columns Symbol + Qty. Found: {list(df.columns)}")

    demat: dict[str, float] = {}
    for _, row in df.iterrows():
        sym = str(row[sym_col]).strip().upper()
        try:
            qty = float(str(row[qty_col]).replace(",", ""))
        except ValueError:
            continue
        if sym and qty >= 0:
            demat[sym] = demat.get(sym, 0) + qty

    q = db.query(PortfolioTransaction)
    if folio_id:
        q = q.filter(PortfolioTransaction.folio_id == folio_id)
    all_txns  = q.all()
    asset_map = {a.id: a.symbol for a in db.query(PortfolioAsset).all()}

    book: dict[str, float] = {}
    for t in all_txns:
        sym = asset_map.get(t.asset_id)
        if not sym:
            continue
        qty = float(t.quantity)
        if t.trans_type in ("Buy", "Bonus", "Transfer_In", "Split"):
            book[sym] = book.get(sym, 0) + qty
        elif t.trans_type in ("Sell", "Transfer_Out"):
            book[sym] = book.get(sym, 0) - qty

    all_symbols = sorted(set(list(demat.keys()) + list(book.keys())))
    rows = []
    for sym in all_symbols:
        d_qty = demat.get(sym, 0.0)
        b_qty = max(book.get(sym, 0.0), 0.0)
        diff  = d_qty - b_qty
        if b_qty < 0.0001 and d_qty < 0.0001:
            continue
        status = "OK" if abs(diff) < 0.01 else ("DEMAT_EXCESS" if diff > 0 else "BOOK_EXCESS")
        rows.append({"symbol": sym, "demat_qty": d_qty, "book_qty": b_qty,
                     "difference": round(diff, 4), "status": status})

    matched    = sum(1 for r in rows if r["status"] == "OK")
    demat_exc  = sum(1 for r in rows if r["status"] == "DEMAT_EXCESS")
    book_exc   = sum(1 for r in rows if r["status"] == "BOOK_EXCESS")

    return {
        "summary": {"total_symbols": len(rows), "matched": matched,
                    "demat_excess": demat_exc, "book_excess": book_exc},
        "rows": sorted(rows, key=lambda r: r["status"]),
    }


# ── Refresh Quotes ────────────────────────────────────────────────────────────

@router.post("/refresh-quotes")
def refresh_quotes(db: Session = Depends(get_db), _=Depends(get_current_user)):
    assets = db.query(PortfolioAsset).all()
    if not assets:
        return {"message": "No portfolio assets to refresh"}

    updated = 0
    failed  = 0
    now     = datetime.utcnow()

    for asset in assets:
        try:
            # Try NSE first, then BSE
            info = None
            for suffix in (".NS", ".BO"):
                try:
                    t = yf.Ticker(f"{asset.symbol}{suffix}")
                    i = t.info
                    if i and (i.get("regularMarketPrice") or i.get("currentPrice")):
                        info = i
                        break
                except Exception:
                    pass

            if not info:
                failed += 1
                continue

            cmp         = float(info.get("currentPrice") or info.get("regularMarketPrice") or 0)
            prev_cls    = float(info.get("regularMarketPreviousClose") or 0)
            day_chg     = round((cmp - prev_cls) / prev_cls * 100, 4) if prev_cls else None
            w52h        = info.get("fiftyTwoWeekHigh")
            w52l        = info.get("fiftyTwoWeekLow")
            w52_chg_raw = info.get("52WeekChange")
            pct_1y      = round(float(w52_chg_raw) * 100, 4) if w52_chg_raw is not None else None
            mc_raw      = info.get("marketCap")
            mc_cr       = round(float(mc_raw) / 10_000_000, 2) if mc_raw else None  # rupees → crores

            existing = db.query(PortfolioQuote).filter(PortfolioQuote.asset_id == asset.id).first()
            if existing:
                existing.cmp            = cmp
                existing.prev_close     = prev_cls or None
                existing.day_change_pct = day_chg
                existing.week52_high    = w52h
                existing.week52_low     = w52l
                existing.market_cap_cr  = mc_cr
                existing.pct_change_1y  = pct_1y
                existing.fetched_at     = now
            else:
                db.add(PortfolioQuote(
                    asset_id=asset.id, cmp=cmp, prev_close=prev_cls or None,
                    day_change_pct=day_chg, week52_high=w52h, week52_low=w52l,
                    market_cap_cr=mc_cr, pct_change_1y=pct_1y,
                    fetched_at=now,
                ))
            updated += 1
        except Exception as e:
            logger.warning(f"Quote refresh failed for {asset.symbol}: {e}")
            failed += 1

    db.commit()
    return {"updated": updated, "failed": failed, "total": len(assets)}


# ── Portfolio Returns (vs Benchmarks) ─────────────────────────────────────────

BENCHMARK_TICKERS = {
    "nifty50":  "^NSEI",
    "nifty500": "^CRSLDX",
    "sensex":   "^BSESN",
    "bse500":   "^BSP500",
}

BENCHMARK_LABELS = {
    "nifty50":  "Nifty 50",
    "nifty500": "Nifty 500",
    "sensex":   "Sensex",
    "bse500":   "BSE 500",
}


def _fy_start(today: date) -> date:
    """Indian FY starts April 1. Returns April 1 of current FY."""
    if today.month >= 4:
        return date(today.year, 4, 1)
    return date(today.year - 1, 4, 1)


def _cutoff_for_years(today: date, years: int) -> date:
    try:
        return today.replace(year=today.year - years)
    except ValueError:
        return today.replace(year=today.year - years, day=28)


def _cutoff_for_months(today: date, months: int) -> date:
    m = today.month - months
    y = today.year
    while m <= 0:
        m += 12
        y -= 1
    try:
        return today.replace(year=y, month=m)
    except ValueError:
        import calendar
        last_day = calendar.monthrange(y, m)[1]
        return today.replace(year=y, month=m, day=last_day)


def _fetch_price_series(symbols: list[str], start: date, end: date) -> dict[str, dict]:
    """
    Fetch historical Close prices for a list of tickers.
    Returns {ticker_str: {date_obj: price}}
    """
    import pandas as pd
    result = {}
    if not symbols:
        return result

    start_str = (start - timedelta(days=7)).strftime("%Y-%m-%d")
    end_str   = (end + timedelta(days=2)).strftime("%Y-%m-%d")

    for sym in symbols:
        prices: dict[date, float] = {}
        # Index symbols (^NSEI, ^BSESN …) are already complete — don't append .NS/.BO
        suffixes = ("",) if sym.startswith("^") else (".NS", ".BO", "")
        for suffix in suffixes:
            ticker_str = f"{sym}{suffix}" if suffix else sym
            try:
                hist = yf.download(ticker_str, start=start_str, end=end_str,
                                   progress=False, auto_adjust=True, multi_level_index=False)
                if hist is not None and not hist.empty and "Close" in hist.columns:
                    for idx, row in hist.iterrows():
                        d = idx.date() if hasattr(idx, "date") else idx
                        if not pd.isna(row["Close"]):
                            prices[d] = float(row["Close"])
                    if prices:
                        break
            except Exception:
                continue
        result[sym] = prices
    return result


def _price_at(price_series: dict, sym: str, target: date) -> float | None:
    """Closest price on-or-before target date."""
    series = price_series.get(sym, {})
    if not series:
        return None
    candidates = [d for d in series if d <= target]
    if not candidates:
        return None
    return series[max(candidates)]


@router.get("/returns")
def portfolio_returns(
    folio_id: Optional[int] = Query(None),
    benchmarks: str = Query("nifty50,sensex"),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    today = date.today()

    # ── Load data ────────────────────────────────────────────────────────────
    if folio_id:
        folios = db.query(Folio).filter(Folio.id == folio_id).all()
    else:
        folios = db.query(Folio).filter(Folio.is_active == True).order_by(Folio.name).all()
    if not folios:
        return {"periods": [], "error": "No folios found"}

    folio_ids = [f.id for f in folios]
    all_txns = (
        db.query(PortfolioTransaction)
        .filter(PortfolioTransaction.folio_id.in_(folio_ids))
        .order_by(PortfolioTransaction.trade_date)
        .all()
    )
    if not all_txns:
        return {"periods": [], "error": "No transactions"}

    asset_map  = {a.id: a for a in db.query(PortfolioAsset).all()}
    quote_map  = {q.asset_id: q for q in db.query(PortfolioQuote).all()}

    # Normalise trade_date to date objects
    def _d(td):
        return td.date() if hasattr(td, "date") else td

    txn_dates = [_d(t.trade_date) for t in all_txns]
    first_txn_date = min(txn_dates)

    # Current portfolio state
    asset_txns: dict[int, list] = {}
    for t in all_txns:
        asset_txns.setdefault(t.asset_id, []).append(t)

    total_current = 0.0
    asset_current: dict[int, dict] = {}
    for asset_id, txns in asset_txns.items():
        asset = asset_map.get(asset_id)
        if not asset:
            continue
        dicts = _to_txn_dicts(txns)
        qty, _, invested = fin.compute_avg_price_and_qty(dicts)
        q   = quote_map.get(asset_id)
        cmp = float(q.cmp) if q and q.cmp else 0.0
        cv  = qty * cmp if cmp else 0.0
        total_current += cv
        asset_current[asset_id] = {
            "symbol": asset.symbol, "qty": qty, "cmp": cmp, "cv": cv,
        }

    # ── Define periods ───────────────────────────────────────────────────────
    fy_start = _fy_start(today)
    period_defs = [
        {"key": "1m",  "label": "1 Month",   "cutoff": _cutoff_for_months(today, 1),  "short": True},
        {"key": "3m",  "label": "3 Months",  "cutoff": _cutoff_for_months(today, 3),  "short": True},
        {"key": "ytd", "label": "FY YTD",    "cutoff": fy_start,                      "short": True},
        {"key": "1y",  "label": "1 Year",    "cutoff": _cutoff_for_years(today, 1),   "short": False},
        {"key": "3y",  "label": "3 Years",   "cutoff": _cutoff_for_years(today, 3),   "short": False},
        {"key": "5y",  "label": "5 Years",   "cutoff": _cutoff_for_years(today, 5),   "short": False},
        {"key": "10y", "label": "10 Years",  "cutoff": _cutoff_for_years(today, 10),  "short": False},
    ]

    # ── Fetch historical prices (batch) ───────────────────────────────────────
    all_symbols = list({v["symbol"] for v in asset_current.values()})
    min_cutoff  = min(p["cutoff"] for p in period_defs)
    price_series = {}
    if all_symbols:
        try:
            price_series = _fetch_price_series(all_symbols, min_cutoff, today)
        except Exception as e:
            logger.warning(f"Price series fetch failed: {e}")

    # Also current prices from quote_map as fallback
    for asset_id, ac in asset_current.items():
        sym = ac["symbol"]
        if sym not in price_series:
            price_series[sym] = {}
        if ac["cmp"]:
            price_series[sym][today] = ac["cmp"]

    # ── Fetch benchmark prices — use DB-managed list ──────────────────────────
    db_benchmarks = db.query(BenchmarkIndex).filter(BenchmarkIndex.is_active == True).all()

    # Build maps from DB
    db_ticker_map = {b.yahoo_symbol: b.label for b in db_benchmarks}
    db_key_map    = {b.yahoo_symbol: b.yahoo_symbol for b in db_benchmarks}

    # Merge with hardcoded BENCHMARK_TICKERS for backward-compat with old ?benchmarks= param
    # If client passes key-based names (nifty50), resolve via BENCHMARK_TICKERS first
    requested = [b.strip() for b in benchmarks.split(",") if b.strip()]
    selected_bm_tickers: list[str] = []
    for req in requested:
        if req in BENCHMARK_TICKERS:
            selected_bm_tickers.append(BENCHMARK_TICKERS[req])
        elif req.startswith("^") or req in db_key_map:
            selected_bm_tickers.append(req)
    # Fall back to all active DB benchmarks if none matched or default passed
    if not selected_bm_tickers or benchmarks == "nifty50,sensex":
        selected_bm_tickers = [b.yahoo_symbol for b in db_benchmarks]

    bm_series: dict[str, dict] = {}
    if selected_bm_tickers:
        try:
            bm_series = _fetch_price_series(selected_bm_tickers, min_cutoff, today)
        except Exception as e:
            logger.warning(f"Benchmark fetch failed: {e}")

    # ── Compute per-period returns ────────────────────────────────────────────
    periods_result = []

    for pdef in period_defs:
        cutoff: date = pdef["cutoff"]
        is_short: bool = pdef["short"]

        insufficient = first_txn_date > cutoff

        # ── Portfolio return ─────────────────────────────────────────────────
        portfolio_return = None
        actual_cutoff = cutoff if not insufficient else first_txn_date

        try:
            # Holdings as of actual_cutoff
            holdings_at_cutoff: dict[int, list] = {}
            for t in all_txns:
                if _d(t.trade_date) <= actual_cutoff:
                    holdings_at_cutoff.setdefault(t.asset_id, []).append(t)

            opening_value = 0.0
            for asset_id, txns in holdings_at_cutoff.items():
                ac = asset_current.get(asset_id)
                if not ac:
                    continue
                dicts = _to_txn_dicts(txns)
                qty, _, _ = fin.compute_avg_price_and_qty(dicts)
                if qty > 0.0001:
                    sym  = ac["symbol"]
                    hist_price = _price_at(price_series, sym, actual_cutoff)
                    # Fallback: current price
                    if hist_price is None:
                        hist_price = ac["cmp"] or 0.0
                    opening_value += qty * hist_price

            # Build cashflows
            cfs = []
            if opening_value > 0.0001:
                cfs.append((actual_cutoff, -opening_value))

            for t in all_txns:
                td = _d(t.trade_date)
                if td > actual_cutoff:
                    amt = float(t.total_amount)
                    if t.trans_type == "Buy":
                        cfs.append((td, -amt))
                    elif t.trans_type == "Sell":
                        cfs.append((td, amt))

            if total_current > 0:
                cfs.append((today, total_current))

            cfs.sort(key=lambda x: x[0])

            if len(cfs) >= 2:
                xirr_val = fin.xirr(cfs)
                if is_short and not insufficient and opening_value > 0:
                    # For short periods show simple absolute %
                    period_days = (today - actual_cutoff).days
                    net_cash_in = sum(-c for _, c in cfs if c < 0)
                    net_cash_out = sum(c for _, c in cfs if c > 0)
                    if net_cash_in > 0:
                        portfolio_return = round((net_cash_out - net_cash_in) / net_cash_in * 100, 2)
                    else:
                        portfolio_return = xirr_val
                else:
                    portfolio_return = xirr_val

        except Exception as e:
            logger.warning(f"Portfolio return calc failed for {pdef['key']}: {e}")

        # ── Benchmark returns ────────────────────────────────────────────────
        bm_returns: dict[str, float | None] = {}
        for ticker in selected_bm_tickers:
            try:
                start_price = _price_at(bm_series, ticker, actual_cutoff)
                end_price   = _price_at(bm_series, ticker, today)
                if start_price and end_price and start_price > 0:
                    if is_short and not insufficient:
                        bm_returns[ticker] = round((end_price - start_price) / start_price * 100, 2)
                    else:
                        bm_returns[ticker] = fin.cagr(start_price, end_price, actual_cutoff)
                else:
                    bm_returns[ticker] = None
            except Exception:
                bm_returns[ticker] = None

        years_actual = round((today - actual_cutoff).days / 365.25, 1)

        periods_result.append({
            "key": pdef["key"],
            "label": pdef["label"],
            "is_short": is_short,
            "insufficient_data": insufficient,
            "actual_years": years_actual,
            "cutoff_date": actual_cutoff.isoformat(),
            "portfolio_return": portfolio_return,
            "benchmarks": bm_returns,
        })

    # Build label map for response: DB first, then legacy hardcoded
    bm_label_map = {b.yahoo_symbol: b.label for b in db_benchmarks}
    for old_key, ticker in BENCHMARK_TICKERS.items():
        if ticker not in bm_label_map:
            bm_label_map[ticker] = BENCHMARK_LABELS.get(old_key, ticker)

    return {
        "periods": periods_result,
        "total_current_value": round(total_current, 2),
        "selected_benchmarks": [
            {"key": t, "label": bm_label_map.get(t, t)} for t in selected_bm_tickers
        ],
        "fy_start": fy_start.isoformat(),
    }


# ── Portfolio Analytics ────────────────────────────────────────────────────────

def _market_cap_band(mc_cr: float | None) -> str:
    if mc_cr is None:
        return "unknown"
    if mc_cr >= 100_000:
        return "large"
    if mc_cr >= 10_000:
        return "mid"
    return "small"


@router.get("/analytics")
def portfolio_analytics(
    folio_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    today = date.today()
    fy_start = _fy_start(today)

    if folio_id:
        folios = db.query(Folio).filter(Folio.id == folio_id).all()
    else:
        folios = db.query(Folio).filter(Folio.is_active == True).order_by(Folio.name).all()
    if not folios:
        return {"error": "No folios found"}

    folio_ids = [f.id for f in folios]
    all_txns = (
        db.query(PortfolioTransaction)
        .filter(PortfolioTransaction.folio_id.in_(folio_ids))
        .order_by(PortfolioTransaction.trade_date)
        .all()
    )
    if not all_txns:
        return {"holdings": [], "summary": {}}

    asset_map = {a.id: a for a in db.query(PortfolioAsset).all()}
    quote_map = {q.asset_id: q for q in db.query(PortfolioQuote).all()}

    def _d(td):
        return td.date() if hasattr(td, "date") else td

    # Group transactions by asset
    asset_txns: dict[int, list] = {}
    for t in all_txns:
        asset_txns.setdefault(t.asset_id, []).append(t)

    holdings = []
    total_current = 0.0
    total_invested = 0.0

    # For churn: buys/sells in last 12m and FY
    fy_buy_amt = fy_sell_amt = 0.0
    y12_buy_amt = y12_sell_amt = 0.0
    y12_cutoff = _cutoff_for_months(today, 12)

    for t in all_txns:
        td = _d(t.trade_date)
        amt = float(t.total_amount)
        if t.trans_type == "Buy":
            if td >= fy_start:
                fy_buy_amt += amt
            if td >= y12_cutoff:
                y12_buy_amt += amt
        elif t.trans_type == "Sell":
            if td >= fy_start:
                fy_sell_amt += amt
            if td >= y12_cutoff:
                y12_sell_amt += amt

    all_holding_days = []

    for asset_id, txns in asset_txns.items():
        asset = asset_map.get(asset_id)
        if not asset:
            continue
        dicts = _to_txn_dicts(txns)
        qty, avg_p, invested = fin.compute_avg_price_and_qty(dicts)
        q = quote_map.get(asset_id)
        cmp = float(q.cmp) if q and q.cmp else 0.0
        cv  = qty * cmp if cmp else 0.0

        if qty < 0.0001:
            # Exited — still count for holding period
            buy_dates = [_d(t.trade_date) for t in txns if t.trans_type == "Buy"]
            sell_dates = [_d(t.trade_date) for t in txns if t.trans_type == "Sell"]
            if buy_dates and sell_dates:
                all_holding_days.append((max(sell_dates) - min(buy_dates)).days)
            continue

        total_current  += cv
        total_invested += invested

        first_buy = min((_d(t.trade_date) for t in txns if t.trans_type == "Buy"), default=None)
        holding_days = (today - first_buy).days if first_buy else 0
        all_holding_days.append(holding_days)

        cfs = fin.build_xirr_cashflows(dicts, cv)
        xirr_val = fin.xirr(cfs) if cfs else None
        cagr_val = fin.cagr(invested, cv, first_buy) if first_buy and invested else None
        unrealised_pct = round((cv - invested) / invested * 100, 2) if invested else None

        mc_cr  = float(q.market_cap_cr) if q and q.market_cap_cr else None
        pct_1y = float(q.pct_change_1y) if q and q.pct_change_1y else None

        holdings.append({
            "asset_id":         asset_id,
            "symbol":           asset.symbol,
            "asset_name":       asset.name,
            "sector":           asset.sector or "Unknown",
            "current_value":    round(cv, 2),
            "total_investment": round(invested, 2),
            "weight_pct":       None,           # filled after total known
            "unrealised_pnl_pct": unrealised_pct,
            "cagr_pct":         cagr_val,
            "xirr_pct":         xirr_val,
            "pct_change_1y":    pct_1y,         # stock's 52-week return
            "holding_days":     holding_days,
            "first_purchase_date": first_buy.isoformat() if first_buy else None,
            "market_cap_cr":    mc_cr,
            "market_cap_band":  _market_cap_band(mc_cr),
        })

    # Fill weights
    for h in holdings:
        h["weight_pct"] = round(h["current_value"] / total_current * 100, 2) if total_current else 0

    # ── Sector / industry aggregations ───────────────────────────────────────
    sector_weights: dict[str, float] = {}
    for h in holdings:
        sector_weights[h["sector"]] = sector_weights.get(h["sector"], 0) + h["weight_pct"]
    sector_chart = [{"name": k, "value": round(v, 2)} for k, v in sorted(sector_weights.items(), key=lambda x: -x[1])]

    # ── Market cap aggregations ──────────────────────────────────────────────
    mc_bands: dict[str, float] = {"large": 0.0, "mid": 0.0, "small": 0.0, "unknown": 0.0}
    for h in holdings:
        mc_bands[h["market_cap_band"]] = mc_bands.get(h["market_cap_band"], 0) + h["weight_pct"]
    mc_chart = [
        {"name": "Large Cap", "value": round(mc_bands["large"], 2)},
        {"name": "Mid Cap",   "value": round(mc_bands["mid"], 2)},
        {"name": "Small Cap", "value": round(mc_bands["small"], 2)},
    ]
    if mc_bands["unknown"] > 0:
        mc_chart.append({"name": "Unknown", "value": round(mc_bands["unknown"], 2)})

    # ── Winners / losers ─────────────────────────────────────────────────────
    def _top_losers(key: str, n: int = 5):
        valid = [h for h in holdings if h.get(key) is not None]
        best  = sorted(valid, key=lambda h: -h[key])[:n]
        worst = sorted(valid, key=lambda h: h[key])[:n]
        return best, worst

    winners_xirr, losers_xirr = _top_losers("xirr_pct")
    winners_1y, losers_1y     = _top_losers("pct_change_1y")

    # Period winners — by CAGR, filtered by holding length
    def _period_wl(min_days: int, max_days: int | None, n: int = 5):
        valid = [h for h in holdings
                 if h.get("cagr_pct") is not None
                 and h["holding_days"] >= min_days
                 and (max_days is None or h["holding_days"] < max_days)]
        return (
            sorted(valid, key=lambda h: -h["cagr_pct"])[:n],
            sorted(valid, key=lambda h: h["cagr_pct"])[:n],
        )

    w1y, l1y = _period_wl(365, 3 * 365)
    w3y, l3y = _period_wl(3 * 365, 5 * 365)
    w5y, l5y = _period_wl(5 * 365, None)

    # ── Summary stats ─────────────────────────────────────────────────────────
    avg_holding = round(sum(all_holding_days) / len(all_holding_days), 0) if all_holding_days else 0
    fy_churn    = round((fy_buy_amt + fy_sell_amt) / total_current * 100, 1) if total_current else None

    return {
        "summary": {
            "total_stocks":       len(holdings),
            "total_value":        round(total_current, 2),
            "total_invested":     round(total_invested, 2),
            "avg_holding_days":   int(avg_holding),
            "fy_churn_pct":       fy_churn,
            "fy_buy_amount":      round(fy_buy_amt, 2),
            "fy_sell_amount":     round(fy_sell_amt, 2),
        },
        "holdings":           holdings,
        "sector_chart":       sector_chart,
        "market_cap_chart":   mc_chart,
        "winners": {
            "since_inception": winners_xirr,
            "by_1y_stock_perf": winners_1y,
            "held_1y_3y": w1y,
            "held_3y_5y": w3y,
            "held_5y_plus": w5y,
        },
        "losers": {
            "since_inception": losers_xirr,
            "by_1y_stock_perf": losers_1y,
            "held_1y_3y": l1y,
            "held_3y_5y": l3y,
            "held_5y_plus": l5y,
        },
    }

# ── Dividends ─────────────────────────────────────────────────────────────────

@router.post("/dividends/sync")
def sync_dividends(db: Session = Depends(get_db), _=Depends(require_fm_or_above)):
    """Incremental sync: delete old Dividend transactions, fetch from Yahoo Finance."""

    # 1. Remove legacy Dividend-type transactions
    deleted_txns = db.query(PortfolioTransaction).filter(
        PortfolioTransaction.trans_type == "Dividend"
    ).delete()
    db.flush()

    # 2. All non-dividend transactions per (folio, asset)
    all_txns = (
        db.query(PortfolioTransaction)
        .order_by(PortfolioTransaction.folio_id, PortfolioTransaction.asset_id, PortfolioTransaction.trade_date)
        .all()
    )
    folio_asset_txns: dict = {}
    for t in all_txns:
        folio_asset_txns.setdefault((t.folio_id, t.asset_id), []).append(t)

    if not folio_asset_txns:
        db.commit()
        return {"deleted_transactions": deleted_txns, "synced": 0, "errors": []}

    # 3. Last synced ex_date per (folio, asset) for incremental
    last_sync: dict = {}
    for d in db.query(PortfolioDividend).all():
        key = (d.folio_id, d.asset_id)
        if key not in last_sync or d.ex_date > last_sync[key]:
            last_sync[key] = d.ex_date

    # 4. Group by asset
    asset_folio_ids: dict = {}
    for (fi, ai) in folio_asset_txns:
        asset_folio_ids.setdefault(ai, set()).add(fi)

    asset_map = {a.id: a for a in db.query(PortfolioAsset).all()}
    synced = 0
    skipped = 0
    errors = []

    for asset_id, folio_ids in asset_folio_ids.items():
        asset = asset_map.get(asset_id)
        if not asset:
            continue

        folio_lasts = [last_sync.get((fi, asset_id)) for fi in folio_ids]
        fetch_start = min(folio_lasts) if all(d is not None for d in folio_lasts) else None

        raw_divs = None
        for suffix in (".NS", ".BO", ""):
            try:
                t = yf.Ticker(f"{asset.symbol}{suffix}")
                d = t.dividends
                if d is not None and not d.empty:
                    raw_divs = d
                    break
            except Exception:
                continue

        if raw_divs is None or raw_divs.empty:
            skipped += 1
            continue

        div_list = []
        for idx, amount in raw_divs.items():
            try:
                ex_dt = idx.date() if hasattr(idx, "date") else idx
                if fetch_start and ex_dt <= fetch_start:
                    continue
                div_list.append((ex_dt, float(amount)))
            except Exception:
                continue

        if not div_list:
            continue

        for folio_id in folio_ids:
            key = (folio_id, asset_id)
            folio_last = last_sync.get(key)
            txns = folio_asset_txns.get(key, [])

            for ex_dt, dps in div_list:
                if folio_last and ex_dt <= folio_last:
                    continue

                qty = 0.0
                for t in txns:
                    if t.trade_date <= ex_dt:
                        if t.trans_type in ("Buy", "Bonus", "Split", "Transfer_In"):
                            qty += float(t.quantity)
                        elif t.trans_type in ("Sell", "Transfer_Out"):
                            qty -= float(t.quantity)
                qty = max(0.0, qty)
                if qty < 0.001:
                    continue

                total = round(qty * dps, 2)
                exists = db.query(PortfolioDividend).filter(
                    PortfolioDividend.folio_id == folio_id,
                    PortfolioDividend.asset_id == asset_id,
                    PortfolioDividend.ex_date == ex_dt,
                ).first()
                if not exists:
                    db.add(PortfolioDividend(
                        folio_id=folio_id, asset_id=asset_id, ex_date=ex_dt,
                        dividend_per_share=round(dps, 6),
                        qty_held=round(qty, 4),
                        total_received=total,
                    ))
                    synced += 1

    db.commit()
    return {"deleted_transactions": deleted_txns, "synced": synced, "skipped_no_data": skipped, "errors": errors}


@router.get("/dividends")
def list_dividends(
    folio_id:   Optional[int]  = Query(None),
    symbol:     Optional[str]  = Query(None),
    from_date:  Optional[date] = Query(None),
    to_date:    Optional[date] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(PortfolioDividend).filter(PortfolioDividend.total_received > 0)
    if folio_id:
        q = q.filter(PortfolioDividend.folio_id == folio_id)
    if from_date:
        q = q.filter(PortfolioDividend.ex_date >= from_date)
    if to_date:
        q = q.filter(PortfolioDividend.ex_date <= to_date)

    divs = q.order_by(PortfolioDividend.ex_date.desc()).all()
    folio_map = {f.id: f.name for f in db.query(Folio).all()}
    asset_map = {a.id: a for a in db.query(PortfolioAsset).all()}

    if symbol:
        sym_up = symbol.upper()
        asset_match = next((a for a in asset_map.values() if a.symbol.upper() == sym_up), None)
        divs = [d for d in divs if d.asset_id == asset_match.id] if asset_match else []

    return [
        {
            "id": d.id,
            "folio_id": d.folio_id,
            "folio_name": folio_map.get(d.folio_id, ""),
            "asset_id": d.asset_id,
            "symbol": asset_map[d.asset_id].symbol if d.asset_id in asset_map else "",
            "asset_name": asset_map[d.asset_id].name if d.asset_id in asset_map else "",
            "ex_date": d.ex_date.isoformat(),
            "dividend_per_share": float(d.dividend_per_share or 0),
            "qty_held": float(d.qty_held or 0),
            "total_received": float(d.total_received or 0),
        }
        for d in divs
    ]


@router.get("/dividends/totals")
def dividend_totals(
    folio_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(PortfolioDividend).filter(PortfolioDividend.total_received > 0)
    if folio_id:
        q = q.filter(PortfolioDividend.folio_id == folio_id)

    divs = q.all()
    asset_map = {a.id: a for a in db.query(PortfolioAsset).all()}
    cutoff = date.today() - timedelta(days=365)

    total_all_time = round(sum(float(d.total_received or 0) for d in divs), 2)
    trailing_12m   = round(sum(float(d.total_received or 0) for d in divs if d.ex_date >= cutoff), 2)

    by_stock: dict = {}
    for d in divs:
        asset = asset_map.get(d.asset_id)
        sym  = asset.symbol if asset else str(d.asset_id)
        name = asset.name   if asset else sym
        if sym not in by_stock:
            by_stock[sym] = {"symbol": sym, "asset_name": name, "total": 0.0, "trailing_12m": 0.0}
        by_stock[sym]["total"] += float(d.total_received or 0)
        if d.ex_date >= cutoff:
            by_stock[sym]["trailing_12m"] += float(d.total_received or 0)

    for s in by_stock.values():
        s["total"] = round(s["total"], 2)
        s["trailing_12m"] = round(s["trailing_12m"], 2)

    last_row = db.query(PortfolioDividend.fetched_at).order_by(PortfolioDividend.fetched_at.desc()).first()
    return {
        "total_all_time": total_all_time,
        "trailing_12m":   trailing_12m,
        "by_stock": sorted(by_stock.values(), key=lambda x: -x["total"]),
        "last_sync": last_row[0].isoformat() if last_row and last_row[0] else None,
    }


@router.delete("/dividends/{dividend_id}", status_code=204)
def delete_dividend(dividend_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    """Delete a dividend record. Inserts a zero-value sentinel so sync never re-fetches this ex_date."""
    rec = db.query(PortfolioDividend).filter(PortfolioDividend.id == dividend_id).first()
    if not rec:
        raise HTTPException(404, "Dividend record not found")
    # Sentinel: keeps the ex_date as a sync boundary but shows nothing in the UI
    sentinel = PortfolioDividend(
        folio_id=rec.folio_id, asset_id=rec.asset_id,
        ex_date=rec.ex_date, dividend_per_share=0, qty_held=0, total_received=0,
    )
    db.delete(rec)
    db.flush()
    # Only add sentinel if no other record exists for this (folio, asset, ex_date)
    existing = db.query(PortfolioDividend).filter(
        PortfolioDividend.folio_id == rec.folio_id,
        PortfolioDividend.asset_id == rec.asset_id,
        PortfolioDividend.ex_date == rec.ex_date,
    ).first()
    if not existing:
        db.add(sentinel)
    db.commit()


# ── Benchmarks ────────────────────────────────────────────────────────────────

DEFAULT_BENCHMARKS = [
    {"label": "Nifty 50",         "yahoo_symbol": "^NSEI"},
    {"label": "Sensex",           "yahoo_symbol": "^BSESN"},
    {"label": "Nifty Midcap 150", "yahoo_symbol": "^NSEMDCP150"},
    {"label": "Nifty Next 50",    "yahoo_symbol": "^NSMIDCP"},
]


@router.get("/benchmarks")
def list_benchmarks(db: Session = Depends(get_db), _=Depends(get_current_user)):
    rows = db.query(BenchmarkIndex).order_by(BenchmarkIndex.label).all()
    # Seed defaults if table is empty
    if not rows:
        for bm in DEFAULT_BENCHMARKS:
            db.add(BenchmarkIndex(label=bm["label"], yahoo_symbol=bm["yahoo_symbol"], is_active=True))
        db.commit()
        rows = db.query(BenchmarkIndex).order_by(BenchmarkIndex.label).all()
    return [{"id": b.id, "label": b.label, "yahoo_symbol": b.yahoo_symbol, "is_active": b.is_active} for b in rows]


@router.post("/benchmarks", status_code=201)
def create_benchmark(body: dict, db: Session = Depends(get_db), _=Depends(get_current_user)):
    label  = (body.get("label") or "").strip()
    symbol = (body.get("yahoo_symbol") or "").strip()
    if not label or not symbol:
        raise HTTPException(400, "label and yahoo_symbol are required")
    existing = db.query(BenchmarkIndex).filter(BenchmarkIndex.yahoo_symbol == symbol).first()
    if existing:
        existing.label = label; existing.is_active = True
        db.commit(); db.refresh(existing)
        return {"id": existing.id, "label": existing.label, "yahoo_symbol": existing.yahoo_symbol, "is_active": existing.is_active}
    bm = BenchmarkIndex(label=label, yahoo_symbol=symbol, is_active=True)
    db.add(bm); db.commit(); db.refresh(bm)
    return {"id": bm.id, "label": bm.label, "yahoo_symbol": bm.yahoo_symbol, "is_active": bm.is_active}


@router.delete("/benchmarks/{bm_id}", status_code=204)
def delete_benchmark(bm_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    bm = db.query(BenchmarkIndex).filter(BenchmarkIndex.id == bm_id).first()
    if not bm:
        raise HTTPException(404, "Benchmark not found")
    db.delete(bm); db.commit()
